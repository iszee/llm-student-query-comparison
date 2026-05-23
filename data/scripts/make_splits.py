"""
make_splits.py
--------------
Produces train.jsonl, test.jsonl, and few_shot_examples.json for LoRA
fine-tuning (Unsloth, A100). Targets Qwen2.5-14B and Gemma latest via
HuggingFace messages format.

Sources (read-only):
  data-collection/manual/data-manual.xlsx               -- 38 human-authored pairs
  data-collection/generated/corrected/corrected-qa.csv  -- 183 human-corrected pairs
  data-collection/generated/generated-qa-combined.csv   -- 2043 AI-generated pairs

Outputs (written to data/):
  data/train.jsonl              -- training set in HF messages format
  data/test.jsonl               -- 50 pairs, human-validated only
  data/few_shot_examples.json   -- 5 held-out pairs for inference-time few-shot
  data/split_stats.json         -- seed, source counts, final totals

Allocation of human-validated pool (corrected + manual):
  First  FEW_SHOT_SIZE pairs  -> few-shot (excluded from train AND test)
  Next   TEST_SIZE pairs      -> test (stratified across sources)
  Rest                        -> train_human

Test/few-shot constraint: never drawn from AI-generated data.

Sampling: stratified by source, seed=42, reproducible.

Usage:
  python data/scripts/make_splits.py

---- Unsloth inference with few-shot ----------------------------------------
  import json
  from pathlib import Path

  few_shot = json.loads(Path("data/few_shot_examples.json").read_text())
  # few_shot is a list of {role, content} turns (user+assistant alternating)

  messages = [
      {"role": "system", "content": SYSTEM_PROMPT},
      *few_shot,                          # prepend examples
      {"role": "user", "content": question},
  ]
  inputs = tokenizer.apply_chat_template(messages, return_tensors="pt")

---- Unsloth training --------------------------------------------------------
  from datasets import load_dataset
  ds = load_dataset("json", data_files={"train": "data/train.jsonl",
                                        "test":  "data/test.jsonl"})
  def fmt(ex):
      return {"text": tokenizer.apply_chat_template(
          ex["messages"], tokenize=False, add_generation_prompt=False)}
  ds = ds.map(fmt)
  # pass ds["train"] to SFTTrainer as usual
"""

import csv
import json
import random
from pathlib import Path

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).parent.parent.parent
DATA_DIR = REPO_ROOT / "data"

MANUAL_XLSX   = REPO_ROOT / "data-collection" / "manual" / "data-manual.xlsx"
CORRECTED_CSV = REPO_ROOT / "data-collection" / "generated" / "corrected" / "corrected-qa.csv"
COMBINED_CSV  = REPO_ROOT / "data-collection" / "generated" / "generated-qa-combined.csv"

TRAIN_JSONL      = DATA_DIR / "train.jsonl"
TEST_JSONL       = DATA_DIR / "test.jsonl"
FEW_SHOT_JSON    = DATA_DIR / "few_shot_examples.json"
STATS_JSON       = DATA_DIR / "split_stats.json"

# ── Config ────────────────────────────────────────────────────────────────────
SEED          = 42
FEW_SHOT_SIZE = 5   # pairs reserved for inference-time few-shot (never in train/test)
TEST_SIZE     = 50  # pairs for evaluation (human-validated only)

SYSTEM_PROMPT = (
    "You are a helpful information assistant for the UQ Bachelor of Information "
    "Technology program. Answer student questions accurately and concisely. "
    "Only provide information you are confident about. "
    "If you are not certain about a specific fact, say so and direct the student "
    "to study.uq.edu.au rather than guessing."
)


# ── Loaders ───────────────────────────────────────────────────────────────────

def load_xlsx(path: Path) -> list[tuple[str, str]]:
    """Load (Question, Answer) pairs from an Excel file."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]
    q_col = headers.index("Question")
    a_col = headers.index("Answer")
    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = str(row[q_col]).strip() if row[q_col] else ""
        a = str(row[a_col]).strip() if row[a_col] else ""
        if q and a:
            pairs.append((q, a))
    return pairs


def load_csv(path: Path) -> list[tuple[str, str]]:
    """Load (Question, Answer) pairs from a CSV with # / Question / Answer columns."""
    pairs = []
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            q = row.get("Question", "").strip()
            a = row.get("Answer", "").strip()
            if q and a:
                pairs.append((q, a))
    return pairs


# ── Formatter ─────────────────────────────────────────────────────────────────

def to_messages(q: str, a: str) -> dict:
    """Wrap a Q/A pair in the HuggingFace messages format (system + user + assistant)."""
    return {
        "messages": [
            {"role": "system",    "content": SYSTEM_PROMPT},
            {"role": "user",      "content": q},
            {"role": "assistant", "content": a},
        ]
    }


def write_jsonl(path: Path, pairs: list[tuple[str, str]]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for q, a in pairs:
            f.write(json.dumps(to_messages(q, a), ensure_ascii=False) + "\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    rng = random.Random(SEED)

    # ── 1. Load ───────────────────────────────────────────────────────────────
    print("Loading sources...")
    manual    = load_xlsx(MANUAL_XLSX)
    corrected = load_csv(CORRECTED_CSV)
    generated = load_csv(COMBINED_CSV)
    print(f"  manual:    {len(manual):4d} pairs")
    print(f"  corrected: {len(corrected):4d} pairs")
    print(f"  generated: {len(generated):4d} pairs")
    print()

    # ── 2. Shuffle human pool once (seed-stable) ──────────────────────────────
    # Allocation order from the shuffled list:
    #   [0 : FEW_SHOT_SIZE]            -> few-shot examples
    #   [FEW_SHOT_SIZE : FEW_SHOT_SIZE + TEST_SIZE]  -> test (stratified)
    #   [FEW_SHOT_SIZE + TEST_SIZE :]  -> train_human

    human_total = len(corrected) + len(manual)
    reserved    = FEW_SHOT_SIZE + TEST_SIZE   # pairs taken from human pool for non-train

    # Stratified proportions for few-shot + test combined
    n_corrected_reserved = round(reserved * len(corrected) / human_total)
    n_manual_reserved    = reserved - n_corrected_reserved

    corrected_shuffled = corrected[:]
    rng.shuffle(corrected_shuffled)
    manual_shuffled = manual[:]
    rng.shuffle(manual_shuffled)

    # Proportional few-shot split
    n_fs_corrected = round(FEW_SHOT_SIZE * len(corrected) / human_total)
    n_fs_manual    = FEW_SHOT_SIZE - n_fs_corrected

    few_shot_pairs = (
        corrected_shuffled[:n_fs_corrected]
        + manual_shuffled[:n_fs_manual]
    )
    assert len(few_shot_pairs) == FEW_SHOT_SIZE

    # Test pairs (next slice)
    n_test_corrected = round(TEST_SIZE * len(corrected) / human_total)
    n_test_manual    = TEST_SIZE - n_test_corrected

    test_pairs = (
        corrected_shuffled[n_fs_corrected : n_fs_corrected + n_test_corrected]
        + manual_shuffled[n_fs_manual : n_fs_manual + n_test_manual]
    )
    assert len(test_pairs) == TEST_SIZE

    # Remaining human pairs for train
    train_human = (
        corrected_shuffled[n_fs_corrected + n_test_corrected:]
        + manual_shuffled[n_fs_manual + n_test_manual:]
    )

    # ── 3. Safety dedup: no few-shot or test question leaks into train ─────────
    reserved_keys = {q.strip().lower() for q, _ in few_shot_pairs + test_pairs}

    train_human_clean = [(q, a) for q, a in train_human  if q.strip().lower() not in reserved_keys]
    generated_clean   = [(q, a) for q, a in generated    if q.strip().lower() not in reserved_keys]

    leaked_human = len(train_human) - len(train_human_clean)
    leaked_gen   = len(generated)   - len(generated_clean)

    # ── 4. Combine and shuffle train ──────────────────────────────────────────
    train_pairs = generated_clean + train_human_clean
    rng.shuffle(train_pairs)

    # ── 5. Write outputs ──────────────────────────────────────────────────────
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    write_jsonl(TEST_JSONL,  test_pairs)
    write_jsonl(TRAIN_JSONL, train_pairs)

    # Few-shot: flat list of alternating user/assistant turns for easy prepending
    few_shot_turns = []
    for q, a in few_shot_pairs:
        few_shot_turns.append({"role": "user",      "content": q})
        few_shot_turns.append({"role": "assistant", "content": a})

    FEW_SHOT_JSON.write_text(
        json.dumps(
            {
                "description": (
                    "Held-out human-validated examples for inference-time few-shot prompting. "
                    "Prepend these turns after the system message and before the user question."
                ),
                "system_prompt": SYSTEM_PROMPT,
                "turns": few_shot_turns,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    # ── 6. Stats ──────────────────────────────────────────────────────────────
    stats = {
        "seed": SEED,
        "system_prompt": SYSTEM_PROMPT,
        "sources": {
            "manual":    len(manual),
            "corrected": len(corrected),
            "generated": len(generated),
        },
        "few_shot": {
            "total":          len(few_shot_pairs),
            "from_corrected": n_fs_corrected,
            "from_manual":    n_fs_manual,
        },
        "test": {
            "total":          len(test_pairs),
            "from_corrected": n_test_corrected,
            "from_manual":    n_test_manual,
        },
        "train": {
            "total":                 len(train_pairs),
            "from_generated":        len(generated_clean),
            "from_human_remaining":  len(train_human_clean),
        },
        "leakage_removed": {
            "from_train_human": leaked_human,
            "from_generated":   leaked_gen,
        },
    }
    STATS_JSON.write_text(json.dumps(stats, indent=2, ensure_ascii=False), encoding="utf-8")

    # ── 7. Summary ────────────────────────────────────────────────────────────
    print("Split complete")
    print(f"  few_shot_examples.json  {len(few_shot_pairs):4d} pairs  "
          f"({n_fs_corrected} corrected + {n_fs_manual} manual)  -- inference only")
    print(f"  test.jsonl              {len(test_pairs):4d} pairs  "
          f"({n_test_corrected} corrected + {n_test_manual} manual)")
    print(f"  train.jsonl             {len(train_pairs):4d} pairs  "
          f"({len(generated_clean)} generated + {len(train_human_clean)} human)")
    if leaked_human or leaked_gen:
        print(f"  Leakage removed: {leaked_human} human train, {leaked_gen} generated")
    else:
        print("  No leakage -- train/test/few-shot are fully disjoint.")
    print()
    print(f"  Output: {DATA_DIR}")
    print(f"  Seed:   {SEED}")


if __name__ == "__main__":
    main()
