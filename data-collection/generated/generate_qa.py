"""
Generate 200 synthetic Q&A pairs for the University Bachelor of Information Technology and related dual degrees.
Sources: the University program pages (study.example.edu, programs-courses.example.edu),
         the University International Guide 2026, manual dataset format.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

QA_PAIRS = [
    # ── A. DOMESTIC ADMISSION & ENTRY REQUIREMENTS (Q1–Q10) ──────────────────
    (
        "What is the minimum ATAR required to apply for the Bachelor of Information Technology at the University?",
        "The minimum adjusted score for Semester 1, 2026 was 81.9, with a median of 87.4 and a highest of 95.45. Meeting the minimum does not guarantee admission. Entry scores are reviewed annually in April, so check study.example.edu for the most current information."
    ),
    (
        "What Year 12 subject prerequisites do I need to apply for the BIT at the University as a domestic student?",
        "You need General English (Units 3 & 4, C minimum) and at least one mathematics subject at Units 3 & 4 with a C minimum. Mathematical Methods or Specialist Mathematics are recommended. Specialist Mathematics provides increased flexibility in later courses, but is not mandatory."
    ),
    (
        "What ATAR do I need for the Bachelors of Engineering (Honours) / Information Technology dual degree?",
        "The minimum adjusted score for Semester 1, 2026 was 84, with a median of 87.25. This program also requires General English, Mathematical Methods, and either Chemistry or Physics (all Units 3 & 4, C minimum). Check study.example.edu as scores are updated annually."
    ),
    (
        "Do all the IT dual degrees at the University require the same ATAR?",
        "Most IT dual degrees at the University require a minimum ATAR of 84. The standalone BIT has a lower minimum of 81.9. Entry scores are reviewed annually and can change, so check the current requirements on study.example.edu for the specific dual degree you are considering."
    ),
    (
        "Is the BIT program at the University a quota program with limited places?",
        "The BIT program is not listed as a quota program. However, entry is competitive and meeting the minimum score does not guarantee a place. Always verify current availability on study.example.edu before applying."
    ),
    (
        "As a domestic student, how do I apply for the BIT at the University - through the state admissions centre or directly?",
        "Domestic undergraduate students apply through the state admissions centre (the state Tertiary Admissions Centre). The state admissions centre code for the Bachelor of Information Technology is 733001. Interstate students also apply through the state admissions centre."
    ),
    (
        "I completed Year 12 in another Australian state. Can I apply for BIT at the University?",
        "Yes. Interstate students apply for BIT through the state admissions centre. Your Year 12 results will be converted to an equivalent entry score. You should confirm that your subjects satisfy the prerequisite requirements and contact the state admissions centre or the University Admissions if you are unsure."
    ),
    (
        "I did not study Specialist Mathematics in Year 12. Can I still apply for the BIT?",
        "Yes. The BIT prerequisite requires at least one of General Mathematics, Mathematical Methods, or Specialist Mathematics at Units 3 & 4 with a C minimum. Specialist Mathematics is not mandatory but is recommended for greater flexibility in later courses."
    ),
    (
        "What is the minimum ATAR for the Bachelors of Business Management / Information Technology?",
        "The minimum adjusted score for Semester 1, 2026 was 84, with a median of 89.2 and a highest of 93.4. Entry scores are reviewed annually, so check study.example.edu for the most current information before applying."
    ),
    (
        "What are the prerequisites and ATAR for the Bachelors of Commerce / IT as a domestic student?",
        "You need General English (Units 3 & 4, C minimum) and Mathematical Methods (Units 3 & 4, C minimum). The minimum ATAR for 2026 entry was 84. Domestic students apply through the state admissions centre using code 711621."
    ),

    # ── B. INTERNATIONAL ADMISSION & ENTRY REQUIREMENTS (Q11–Q24) ────────────
    (
        "As an international student, what IELTS score do I need for the BIT?",
        "The BIT requires 'Minimum' English proficiency: IELTS overall 6.5 with no sub-band below 6. Equivalent alternatives are TOEFL iBT 87 overall (listening 19, reading 19, writing 21, speaking 19), or PTE Academic overall 64 with at least 60 in all sub-bands."
    ),
    (
        "I have a Chinese GaoKao result. Can I use it to apply for BIT at the University?",
        "For Chinese students, successful completion of the first year of a bachelor's degree at a Chinese university is typically the state Year 12 equivalent accepted by the University. You should contact the University International Admissions at anon@example.edu to confirm your specific eligibility."
    ),
    (
        "I completed the All India Senior School Certificate (CBSE). Can I apply for BIT?",
        "Yes. The CBSE is accepted as a the state Year 12 equivalent. For English proficiency, a grade of 65 per cent or better in English from CBSE satisfies the Minimum English requirement. You should also ensure you meet the mathematics prerequisite."
    ),
    (
        "What IB score is needed for the IT dual degrees such as Commerce/IT or Business Management/IT?",
        "A minimum International Baccalaureate (IB) score of 32 is required for most IT dual degrees at the University, including BCom/IT and BBusMgt/IT. You should also ensure your IB subjects meet the English and mathematics prerequisites for the specific program."
    ),
    (
        "Do I need to take a separate English language test if my previous study was conducted entirely in English?",
        "Not necessarily. If you completed senior secondary schooling or at least one year of full-time post-secondary study at an institution in a designated English-speaking location, you may meet the University's English requirements without a separate test. Designated locations include Australia, Canada, Ireland, New Zealand, the UK, and the USA. Check study.example.edu for the full list."
    ),
    (
        "I have a Malaysian STPM result. Am I eligible to apply for the BIT at the University?",
        "The Malaysian STPM or Unified Examination Certificate (UEC) is accepted as a the state Year 12 equivalent. The minimum entry score for the standalone BIT was 81.9 ATAR equivalent, and 84 for dual degrees. Contact the University International Admissions for guidance on converting your specific result."
    ),
    (
        "When is the international application deadline for Semester 1, 2026 entry into BIT?",
        "The recommended international application deadline for Semester 1 is 30 November of the previous year. This allows enough time to apply for a student visa. Some programs may have earlier closing dates, so check study.example.edu well in advance."
    ),
    (
        "When is the international application deadline for Semester 2 entry into BIT at the University?",
        "The recommended application deadline for international students applying for Semester 2 is 31 May of the same year. Apply well before this date to allow time for visa processing."
    ),
    (
        "Is there an application fee for international students applying to the University?",
        "Yes. A non-refundable application fee of AUD $150 applies to all international applications, regardless of the submission method."
    ),
    (
        "I am an international student studying Year 12 in Australia. How do I apply for the BIT?",
        "If you are an international student currently studying Year 12 in Australia, you can apply through the state admissions centre, the same as domestic students. You may also be eligible for Adjustment factors. Contact anon@example.edu or check study.example.edu/information-resources/high-school-students for specific guidance."
    ),
    (
        "Can I receive credit for previous university study when entering BIT as an international student?",
        "Potentially, yes. If you have undertaken prior formal learning, you may receive credit for some courses in your BIT program. You will still need to complete the standard total number of units to graduate. Check your credit eligibility at study.example.edu/admissions/undergraduate/check-credit-eligibility."
    ),
    (
        "What documents do I need to submit with my the University international application?",
        "You need to provide certified copies of academic transcripts and award certificates, a grading scale for any tertiary studies, certified official translations for non-English documents, a copy of your passport, evidence of English language proficiency, and any program-specific documents. You may also need to complete a Genuine Student Assessment."
    ),
    (
        "I have Singapore GCE A Levels. What are the entry requirements for the BIT at the University?",
        "Singapore-Cambridge GCE A Levels are accepted as the state Year 12 equivalent. For English proficiency, you need a grade of C or better in the General Paper or Knowledge and Inquiry subject, or grade C in H2 English Language and Linguistics or Literature in English. You also need to satisfy the mathematics prerequisite and meet the minimum entry score."
    ),
    (
        "What pathway is available at the University if I do not meet the BIT entry requirements?",
        "If you do not meet the academic or English language entry requirements for BIT, you may be able to enrol in a University College Foundation Program or other academic pathway program. Upon successful completion with the required grades, you can apply for entry into the University degree programs. More information is available at study.example.edu or pages 20–27 of the University International Guide 2026."
    ),

    # ── C. PROGRAM STRUCTURE & DURATION (Q25–Q40) ────────────────────────────
    (
        "How long does the Bachelor of Information Technology take to complete full-time?",
        "The Bachelor of Information Technology is a 3-year full-time program, or part-time equivalent. It is offered at the main campus with intakes in Semester 1 (23 February 2026) and Semester 2 (27 July 2026)."
    ),
    (
        "How many units do I need to complete to graduate from the BIT?",
        "A standard full-time study load is 8 units per semester (16 units per year). For the 3-year BIT, you need to complete 48 units in total. Check Programs and Courses for the exact unit requirements for your major and commencement year."
    ),
    (
        "Can I study the BIT part-time as a domestic student?",
        "Yes, the BIT can be studied part-time by domestic students. Part-time study means less than 75% of the standard full-time load (i.e. fewer than 6 units per semester). Your degree will take longer than 3 years to complete if studied part-time, so plan your course sequence carefully."
    ),
    (
        "Is the Bachelor of Information Technology available for Semester 2 intake?",
        "Yes. The BIT is available for both Semester 1 (commencing 23 February 2026) and Semester 2 (commencing 27 July 2026) intakes at the main campus."
    ),
    (
        "How long does the Bachelors of Engineering (Honours) / IT dual degree take to complete?",
        "The Bachelors of Engineering (Honours) / Information Technology takes 5.5 years full-time. This is longer than other dual degrees because of the integrated honours component within the Engineering degree."
    ),
    (
        "How long does the Bachelors of Information Technology / Arts take to complete?",
        "The Bachelors of Information Technology / Arts is a 4-year full-time program, available for both Semester 1 and Semester 2 intakes at the main campus."
    ),
    (
        "How long does the Bachelors of Business Management / IT take to complete?",
        "The Bachelors of Business Management / Information Technology is a 4-year full-time program, with intakes in both Semester 1 and Semester 2 at the main campus."
    ),
    (
        "Can I add a Concurrent Diploma in Languages alongside my BIT?",
        "Yes. the University offers a Concurrent Diploma in Languages that can be studied alongside most bachelor's degrees, including BIT. It covers 9 languages, is made up of 16 units, and can be applied for after you are enrolled in a full-time the University program."
    ),
    (
        "Which campus is the Bachelor of Information Technology delivered at?",
        "The BIT and all its dual degree combinations are delivered at the main campus, the University's main campus located 7km from the city's city centre."
    ),
    (
        "Does the BIT include an integrated honours year?",
        "No. The standalone Bachelor of Information Technology (3 years) does not include an integrated honours year. After completing the BIT with a sufficient GPA (5.0/7.0 minimum), you can apply separately for the Bachelor of Information Technology (Honours), a one-year program with code 9001."
    ),
    (
        "What is the standard full-time study load per semester for BIT students?",
        "The standard full-time study load is 8 units per semester (16 units per year). This typically means enrolling in 4 courses per semester."
    ),
    (
        "What topics do BIT students study in the program according to the University International Guide 2026?",
        "According to the University International Guide 2026, BIT students study Human-Computer Interaction, Social and Mobile Computing, Programming in the Large, and Artificial Intelligence. The program offers majors in Software Design, Software Information Systems, User Experience Design, and a Computer Systems minor."
    ),
    (
        "What majors are available in the Bachelor of Information Technology at the University?",
        "The BIT offers three majors: Software Information Systems, Solution Architecture (also referred to as Software Design), and User Experience Design. There is also a Computer Systems minor. Confirm the current major list on Programs and Courses before making your selection."
    ),
    (
        "What are the main areas of study in the IT/Design dual degree?",
        "The IT/Design dual degree focuses on combining design principles with technical expertise. Specialisation options include User Experience Design, Software Design, and Software Information Systems. The program is described as developing skills across 'people, code and data'."
    ),
    (
        "What is the CRICOS code for the Bachelor of Information Technology?",
        "The CRICOS code for the Bachelor of Information Technology is CRIC001. CRICOS registration confirms that a program is approved for international students on a student visa."
    ),
    (
        "Is there a Semester 2 intake for the Bachelors of Engineering (Honours) / IT dual degree?",
        "Yes. The Bachelors of Engineering (Honours) / Information Technology is available for both Semester 1 (23 February 2026) and Semester 2 (27 July 2026) intakes."
    ),

    # ── D. MAJORS & SPECIALISATIONS (Q41–Q54) ────────────────────────────────
    (
        "What is the Software Information Systems major in BIT focused on?",
        "The Software Information Systems major focuses on databases, information systems, and software development. It suits students interested in enterprise systems, data management, and software engineering. Check Programs and Courses for the specific course list and study plan for your commencement year."
    ),
    (
        "What does the Solution Architecture major in the BIT cover?",
        "The Solution Architecture major (also referred to as Software Design in some program materials) focuses on designing and building large-scale software solutions, covering software engineering, system design, and architecture principles. Check Programs and Courses for the specific courses required."
    ),
    (
        "I am interested in UX design and human-centred technology. Which BIT major should I choose?",
        "The User Experience Design (UXD) major is the most relevant choice for interface design and human-centred technology. It is available in the BIT and in some dual degrees. Review the full major requirements and course sequence on Programs and Courses before finalising your enrolment."
    ),
    (
        "Can I complete a User Experience Design major in the Bachelors of Business Management / IT?",
        "Based on available program information, the BBusMgt/IT dual degree includes IT major options including Software Information Systems, Solution Architecture, and User Experience Design. Verify the current options on Programs and Courses for your specific commencement year."
    ),
    (
        "What IT majors are available in the Bachelors of Commerce / IT dual degree?",
        "The Commerce component of the BCom/IT offers majors in Accounting, Business Analytics, Business Information Systems, and Finance. The IT component contributes its own core and major requirements. Check Programs and Courses for the full structure relevant to your commencement year."
    ),
    (
        "Does the BIT/Arts dual degree include an IT major?",
        "Yes. The IT component of the BIT/Arts dual degree includes IT core and major requirements from the Bachelor of Information Technology. The Arts component offers over 45 areas of study. Check the specific program requirements on Programs and Courses for your commencement year."
    ),
    (
        "Can I choose my BIT major after I start the degree?",
        "At the University, you typically do not need to declare your major before starting. However, you should plan course selections with a specific major in mind from the beginning, because major courses have prerequisites and a set sequence. Choosing courses without a plan may delay your graduation."
    ),
    (
        "Is there a Computer Systems minor in the BIT?",
        "The University International Guide 2026 lists Computer Systems as a minor option in the BIT program. A minor is a smaller specialisation within your degree. Check Programs and Courses for the specific course requirements and whether this minor is available in your commencement year."
    ),
    (
        "Is Artificial Intelligence a major in the Bachelor of Information Technology?",
        "Artificial Intelligence is not listed as a standalone major in the BIT. However, AI content is covered within the BIT curriculum. If you want a dedicated AI major, consider the Bachelor of Computer Science, which offers an Artificial Intelligence major, or check BIT elective options for AI-related courses."
    ),
    (
        "What does the Software Design major focus on in the IT/Design dual degree?",
        "The Software Design major in the IT/Design dual degree focuses on combining software development with design thinking. It covers the intersection of code, visual design, and user experience. Review the current course list on Programs and Courses to understand the exact requirements."
    ),
    (
        "How many majors do I need to complete in a BIT dual degree?",
        "In a dual degree, you complete requirements for both component degrees simultaneously. The IT component has its own major requirements, and the other degree has its own. Do not assume the single-degree BIT major requirements apply - follow the study plan specific to your dual degree."
    ),
    (
        "Can I change my BIT major after completing some courses?",
        "You may be able to change your major, but you should not assume all completed courses will count towards the new major. Whether they count depends on how they fit the new major, BIT core requirements, or elective categories. Seek academic advice from the School of the School before making the change."
    ),
    (
        "What BIT elective options are available if I want to explore topics outside my major?",
        "The BIT includes general program electives and free electives that allow you to broaden your studies. Specific electives vary by semester. Check Programs and Courses and the course offerings for the relevant semester before finalising enrolment."
    ),
    (
        "What career paths does the Human Movement and Nutrition Sciences / IT dual degree lead to?",
        "Graduates may pursue roles in health technology, sports analytics, software development in health sectors, system architecture, product design, and digital health. The combination is particularly suited to careers at the intersection of technology and health or sport science."
    ),

    # ── E. SPECIFIC COURSES & STUDY PLANS (Q55–Q70) ──────────────────────────
    (
        "What is CORE1001 and what does it cover?",
        "CORE1001 is Introduction to Software Engineering, typically the first programming course in the BIT and Computer Science programs. It introduces fundamental programming concepts using Python and foundational software engineering practices. No prior programming experience is required as a prerequisite."
    ),
    (
        "What is INFO1200 and why is it important in the BIT?",
        "INFO1200 is Introduction to Information Systems, a foundational first-year course covering database concepts, information management, and the role of information systems in organisations. It is a prerequisite for several later courses in the Software Information Systems major, so completing it early is important."
    ),
    (
        "What is DSGN1400 and what does it focus on?",
        "DSGN1400 is Introduction to Web Design, a first-year course covering the fundamentals of web technologies, HTML, CSS, and user interface design. It is typically a foundational course for the User Experience Design major and the DSGN-series courses."
    ),
    (
        "What is MATH1061, and do BIT students need to take it?",
        "MATH1061 is Discrete Mathematics, covering logic, sets, combinatorics, graph theory, and Boolean algebra. It is a core mathematics course for BIT students, providing the mathematical foundation for algorithm analysis and theoretical computer science. Check your study plan to confirm the recommended semester."
    ),
    (
        "What is DSGN3801 and when should I take it?",
        "DSGN3801 is a capstone project course in the DSGN series, typically taken in the final year of the BIT program. It involves a substantial team project combining design and technical skills. It is a core graduation requirement for BIT students following the DSGN pathway."
    ),
    (
        "Can I overload (take more than 8 units in a semester) in the BIT program?",
        "Overloading beyond the standard 8-unit semester load requires approval. You should contact the School of the School or the Faculty to request permission. Overloading without approval is not permitted and the extra courses may not count towards your program requirements."
    ),
    (
        "What is the difference between a core course, a program elective, and a general elective in the BIT?",
        "Core courses are mandatory for all students and must be completed to graduate. Program electives allow you to choose from a set list of approved courses. General electives give more freedom to take courses from across the university. Your major also has its own required courses. Check Programs and Courses for the breakdown in your commencement year."
    ),
    (
        "Can I take a course from another faculty as a general elective in my BIT program?",
        "Yes, the BIT typically allows some general elective units from other faculties. You should confirm the course is available to non-program students, that you meet any prerequisites, and that it counts as an elective in your BIT structure. Seek academic advice if unsure."
    ),
    (
        "What is CORE2002 and is it a prerequisite for advanced BIT courses?",
        "CORE2002 is Programming in the Large, a second-year course building on CORE1001. It covers object-oriented design, software architecture, and large-scale software development. It is typically a prerequisite for several Year 2 and Year 3 CORE courses in the BIT program."
    ),
    (
        "Is it possible to take a Summer Semester course as a BIT student?",
        "the University offers a Summer Semester running from approximately December to January. Not all courses are offered in Summer Semester, and availability varies by year. Check course availability in Programs and Courses and confirm your timetable before enrolling in Summer Semester."
    ),
    (
        "What first-semester courses should I take as a Semester 2 commencing BIT student?",
        "Semester 2 commencing students should follow the study plan specifically designed for their commencement semester - not the Semester 1 plan, which has a different course sequence. Use the study plan for your specific major and commencement semester available from the School or Programs and Courses."
    ),
    (
        "I passed CORE1001 with a grade of 4 out of 7. Can I continue to second-year CORE courses?",
        "A grade of 4 is a pass on the University's 7-point scale. In most cases, you can progress to second-year CORE courses if you have passed CORE1001 and meet other listed prerequisites. Confirm whether the next course specifies a minimum grade requirement beyond a simple pass."
    ),
    (
        "Do BIT students need to complete a capstone or final-year project to graduate?",
        "Yes. The BIT typically requires a capstone project in the final year. For students following the DSGN pathway, this is usually DSGN3801. The specific capstone requirement depends on your major and commencement year. Check your program requirements on Programs and Courses."
    ),
    (
        "What is INFO2200 and what are its prerequisites?",
        "INFO2200 is Relational Database Systems, a second-year course covering database design, SQL, and relational theory. It typically has INFO1200 as a prerequisite. It is a key course for students in the Software Information Systems major."
    ),
    (
        "Where can I find the official study plan for my BIT major?",
        "Study plans - the suggested semester-by-semester course sequence - are available from the School of Computing (the School) website at cs.example.edu and through Programs and Courses. Use the plan that matches your commencement year and semester."
    ),
    (
        "What is the role of DSGN courses in the BIT program?",
        "DSGN (Design and Computing) courses form the design stream of the BIT. They cover web design, interaction design, team-based design projects, and human-centred computing. DSGN courses are particularly central to the User Experience Design major but also appear in the core curriculum for other majors."
    ),

    # ── F. DUAL DEGREES - SPECIFIC PROGRAMS (Q71–Q95) ────────────────────────
    (
        "What is the difference between the standalone BIT and the Bachelors of Business Management / IT dual degree?",
        "The standalone BIT takes 3 years and focuses entirely on IT, while the BBusMgt/IT dual degree takes 4 years and combines IT with business management, awarding two degrees. The dual degree has a higher minimum ATAR (84 vs 81.9) and higher domestic fees (~$13,265 vs ~$9,730 per year)."
    ),
    (
        "Do I need to choose my engineering specialisation before starting the BE(Hons)/IT dual degree?",
        "The Engineering (Honours)/IT program allows you to explore different engineering specialisations in your first year before committing. Available engineering majors include Chemical, Civil, Electrical, Mechanical, and Mechatronic Engineering. Check Programs and Courses for when you must declare your engineering major."
    ),
    (
        "Does the Bachelors of Engineering (Honours) / IT dual degree include an honours component?",
        "Yes. The Engineering component is a Bachelor of Engineering (Honours), which has an integrated honours component. This is why the combined program takes 5.5 years rather than 4 years typical of most dual degrees."
    ),
    (
        "Which engineering specialisation in the BE(Hons)/IT dual degree has the most overlap with IT?",
        "Electrical Engineering has strong technical overlap with IT topics such as digital systems and communications. Software Engineering, if available as a specialisation, is directly aligned. Review the specific major descriptions in Programs and Courses alongside your career goals to make the best choice."
    ),
    (
        "What is the state admissions centre code for the Bachelors of Engineering (Honours) / Information Technology?",
        "The state admissions centre code for the Bachelors of Engineering (Honours) / Information Technology is 717701."
    ),
    (
        "What is the state admissions centre code for the Bachelors of Business Management / Information Technology?",
        "The state admissions centre code for the Bachelors of Business Management / Information Technology is 710401."
    ),
    (
        "What is the state admissions centre code for the Bachelors of Commerce / Information Technology?",
        "The state admissions centre code for the Bachelors of Commerce / Information Technology is 711621."
    ),
    (
        "What is the state admissions centre code for the Bachelors of Information Technology / Arts?",
        "The state admissions centre code for the Bachelors of Information Technology / Arts is 733201."
    ),
    (
        "What is the state admissions centre code for the Bachelors of Information Technology / Design?",
        "The state admissions centre code for the Bachelors of Information Technology / Design is 733310."
    ),
    (
        "What is the state admissions centre code for the Bachelors of Human Movement and Nutrition Sciences / Information Technology?",
        "The state admissions centre code for the Bachelors of Human Movement and Nutrition Sciences / Information Technology is 720802."
    ),
    (
        "Should I choose BCom/IT or BBusMgt/IT if I am interested in technology and business?",
        "Both are 4-year dual degrees with a minimum ATAR of 84. BCom/IT is more finance and accounting oriented (majors: Accounting, Business Analytics, Business Information Systems, Finance). BBusMgt/IT is broader in management, covering HR, Marketing, International Business, and Entrepreneurship. Your career goals should guide the choice."
    ),
    (
        "Can I choose Psychology as my Arts major in the Bachelors of Information Technology / Arts dual degree?",
        "The Arts component of the BIT/Arts dual degree offers over 45 areas of study, and Psychology is among them. Check the current subject availability for your enrolment year and confirm the Psychology major fits within the Arts degree requirements on Programs and Courses."
    ),
    (
        "Does the IT/Design dual degree take longer than the standalone BIT?",
        "Yes. The standalone BIT takes 3 years, while the Bachelors of Information Technology / Design takes 4 years because you are completing two degrees simultaneously in a combined structure."
    ),
    (
        "What careers can I pursue with a Bachelors of Commerce / IT dual degree?",
        "BCom/IT graduates pursue roles including business systems analyst, software architect, purchasing officer, technology consultant, and system administrator. Salaries for these roles range from approximately AUD $120,000 to $190,000 depending on position and experience."
    ),
    (
        "What career roles are typical for Bachelors of Business Management / IT graduates?",
        "BBusMgt/IT graduates can pursue roles including system architect, software developer, cybersecurity specialist, business analyst, and HR manager. Starting salaries range from approximately $72K–$78K for computing graduates to $110K–$125K for business systems analyst roles."
    ),
    (
        "Can I choose Software Information Systems as my IT major in the BBusMgt/IT dual degree?",
        "Based on available program information, the BBusMgt/IT dual degree includes IT major options such as Software Information Systems, Solution Architecture, and User Experience Design. Verify the current major options on Programs and Courses for your commencement year."
    ),
    (
        "Is the Bachelors of Human Movement and Nutrition Sciences / IT available for Semester 2 intake?",
        "Yes. The Bachelors of Human Movement and Nutrition Sciences / Information Technology is available for both Semester 1 (23 February 2026) and Semester 2 (27 July 2026) at the main campus."
    ),
    (
        "Can international students apply for the Bachelors of Engineering (Honours) / IT dual degree?",
        "Yes. International students can apply for the BE(Hons)/IT. The English requirement is 'Minimum' IELTS 6.5. Program code is 9009, CRICOS CRIC003. International tuition fees are approximately AUD $58,056 per year for 2026."
    ),
    (
        "For a dual degree with IT, do I need to meet entry requirements for both component degrees separately?",
        "For most dual degrees, a single set of entry requirements applies to the combined program. However, some programs may have prerequisites drawn from both component degrees. Always check the specific entry requirements listed for your dual degree on study.example.edu."
    ),
    (
        "Can I transfer from a dual degree back to the standalone Bachelor of Information Technology?",
        "A transfer may be possible, but it depends on which courses you have completed and how they satisfy the BIT program requirements. You should seek academic advice from the School of the School or your faculty before making any change, as the process is not automatic."
    ),
    (
        "What is the total indicative fee for the BE(Hons)/IT dual degree for an international student?",
        "The indicative total fee for the 5.5-year Bachelors of Engineering (Honours) / IT is approximately AUD $418,909, based on the 2026 first-year fee of $58,056. This is indicative only and fees are reviewed annually."
    ),
    (
        "What is the total indicative fee for the BBusMgt/IT dual degree for an international student?",
        "The indicative total fee for the 4-year Bachelors of Business Management / IT is approximately AUD $282,910, based on the 2026 first-year fee of $58,056. This is indicative only."
    ),
    (
        "What is the total indicative fee for the BCom/IT dual degree for an international student?",
        "The indicative total fee for the 4-year Bachelors of Commerce / IT is approximately AUD $282,910, based on the 2026 first-year fee of $58,056. Fees are reviewed annually."
    ),
    (
        "What is the total indicative fee for the standalone BIT for an international student?",
        "Based on the University International Guide 2026, the indicative total fee for the 3-year Bachelor of Information Technology is approximately AUD $201,774, based on the 2026 first-year fee of $58,056. Fees are reviewed annually."
    ),
    (
        "What is the total indicative fee for the BIT/Arts dual degree for an international student?",
        "The indicative total fee for the 4-year Bachelors of Information Technology / Arts is approximately AUD $282,910, based on the 2026 first-year fee of $58,056. This is indicative only."
    ),

    # ── G. FEES & FINANCIAL MATTERS (Q96–Q113) ───────────────────────────────
    (
        "What are the approximate annual tuition fees for a domestic BIT student in 2026?",
        "For domestic Commonwealth Supported students, the approximate annual fee for the Bachelor of Information Technology is AUD $9,730 based on a standard full-time load of 16 units. Domestic students may be eligible for HECS-HELP to defer their fees."
    ),
    (
        "What is HECS-HELP, and can BIT domestic students use it?",
        "HECS-HELP is an Australian Government loan scheme that allows eligible domestic students in Commonwealth Supported Places to defer their tuition fees until they earn above the repayment threshold. Domestic BIT students offered a CSP are eligible. You must apply for CSP status through the state admissions centre."
    ),
    (
        "What are the domestic tuition fees for the Bachelors of Business Management / IT?",
        "The approximate annual domestic tuition fee for the Bachelors of Business Management / IT is AUD $13,265, based on 16 units per year for 2026. This is higher than the standalone BIT because business management courses fall into a different fee band."
    ),
    (
        "What are the domestic tuition fees for the Bachelors of Human Movement and Nutrition Sciences / IT?",
        "The approximate annual domestic tuition fee for the Bachelors of Human Movement and Nutrition Sciences / IT is AUD $13,470 per year, indicative for 2026 commencement."
    ),
    (
        "What are the approximate domestic tuition fees for the Bachelors of Engineering (Honours) / IT?",
        "The approximate annual domestic tuition fee for the Bachelors of Engineering (Honours) / IT is AUD $8,285 per year (based on 16 units), indicative for 2026. Fees are reviewed annually."
    ),
    (
        "Do international BIT students need Overseas Student Health Cover (OSHC)?",
        "Yes. International students on a student visa must purchase and maintain OSHC as a visa condition. OSHC covers basic medical and hospital care from the day you arrive in Australia. the University's preferred OSHC provider is Allianz Care. You can arrange this when you accept your offer and pay your tuition fee deposit."
    ),
    (
        "What is the Student Services and Amenities Fee (SSAF), and do BIT students pay it?",
        "The SSAF is charged each semester for non-academic services. In 2025, the annual SSAF was capped at $365. All enrolled students, including BIT students, pay this fee. It covers student welfare, recreational facilities, and student organisations."
    ),
    (
        "How much money do I need to show when applying for a student visa to study at the University?",
        "From 10 May 2024, student visa applicants must demonstrate a minimum of AUD $29,710 to cover living costs for one year in Australia. You also need to budget for tuition fees, establishment costs, and OSHC on top of this amount."
    ),
    (
        "What is the approximate monthly cost of living for a student in the city studying at the University?",
        "Based on the University International Guide 2026, estimated monthly costs for a student living off-campus in student accommodation range from approximately AUD $1,567 to $3,991, depending on lifestyle. On-campus residential college costs are estimated at approximately $2,463 to $3,463 per month (catered meals included)."
    ),
    (
        "Are there scholarships available for international IT students at the University?",
        "the University offers competitive scholarships covering criteria such as academic excellence, sporting excellence, and specific study areas. The Australian Government also provides Australia Awards scholarships for international students. Check eligibility at scholarships.example.edu. You can indicate interest in scholarship consideration on your application form - no separate application is needed."
    ),
    (
        "Can I work part-time while studying the BIT at the University on a student visa?",
        "Yes. International students on a student visa can work up to 48 hours per fortnight during university semesters (effective 1 July 2023) and full-time during holidays. You should not expect part-time casual work to cover tuition fees or full living expenses. Check your visa conditions for any restrictions."
    ),
    (
        "What is the University fee refund policy if I withdraw from the BIT program?",
        "If you cancel your enrolment by the census date, fees can be refunded in full for continuing students; commencing students may incur an administrative charge. After the census date, refunds are only considered in special circumstances. Refer to the Student Refund Procedures at policies.example.edu for full details."
    ),
    (
        "If my citizenship or residency status changes during my degree, does my tuition fee change?",
        "Yes. If your visa or citizenship status changes, your fee status may change. Undergraduate students must apply for a Commonwealth Supported Place (CSP) through the state admissions centre. You should contact the University administration promptly if your residency status changes during your studies."
    ),
    (
        "Can I access student loans or financial aid as an international BIT student from North America?",
        "the University can assist students from North America (and some European and Latin American countries) in accessing student loan programs offered in their home countries. These loans can help cover part of your educational expenses. More information is at study.example.edu/admissions/financial-aid-international-students."
    ),
    (
        "What is an indicative total fee for a University IT program, and is it fixed for the whole degree?",
        "An indicative total fee is an estimate of the full tuition cost, calculated from the current first-year fee. Tuition fees at the University are reviewed annually and can change each year. The fees you pay in later years may differ from the first-year fee. Check study.example.edu each year for the most current fee."
    ),
    (
        "Are there fee differences between the standalone BIT and IT dual degree programs for domestic students?",
        "Yes. Different disciplines fall into different fee bands. The standalone BIT is approximately AUD $9,730/year for domestic students, while BCom/IT is approximately $12,275/year and Human Movement/IT approximately $13,470/year because each component is charged at the fee band for that discipline."
    ),
    (
        "What financial support does the University offer for students experiencing financial hardship?",
        "the University offers financial support through Student Central, including emergency financial assistance and budget advice. Domestic students may be eligible for Youth Allowance, Austudy, or Abstudy through Centrelink. International students should seek advice from the University International student services about scholarship options and work rights."
    ),
    (
        "Can domestic students in IT dual degrees access HECS-HELP?",
        "Yes. Domestic students in Commonwealth Supported Places across all the University undergraduate programs, including dual degrees, can access HECS-HELP. Apply for a CSP through the state admissions centre. Note that the fee per unit varies by discipline, so courses in different disciplines within your dual degree will be charged at different rates."
    ),

    # ── H. CAMPUS LIFE, SUPPORT & ACCOMMODATION (Q114–Q128) ──────────────────
    (
        "What accommodation options are available at the University for BIT students?",
        "the University offers on-campus accommodation through 10 residential colleges (including a college, a college, Emmanuel, Grace, International House, King's, St John's, a college, The Women's College, and Union College), and a new student residence, the University's newest 610-room student residence at the main campus. Off-campus options include purpose-built student accommodation and private housing supported by the University Accommodation team."
    ),
    (
        "Is O-Week attendance compulsory for international BIT students?",
        "Yes. O-Week orientation is compulsory for all international students. During O-Week you find out how to enrol, receive your student ID card, and learn about campus services and facilities. It is held the week before semester starts. For 2026, Semester 1 O-Week runs from 16–20 February."
    ),
    (
        "What healthcare facilities are available at the main campus for BIT students?",
        "The main campus has a University Health Care clinic, the University Dental clinic, and Campus Pharmacy on campus. There is also a safety escort service, a Safety Bus running from 6pm, and the University SafeZone app connecting students to security officers or emergency services."
    ),
    (
        "Does the University offer a free airport pickup service for new international students?",
        "Yes. the University provides a complimentary airport shuttle service for new international students arriving at the city International Airport during peak arrival times. The International Student Welcome Crew will greet you and transport you to key locations around the city. More information is available at my.example.edu."
    ),
    (
        "What is the an academic preparation program at the University?",
        "the program is a program held before semester starts, designed to help new students - particularly international students - transition into university academic life. It covers study skills, university systems, and academic expectations. It is held before semester starts and is available through Student Support and Wellbeing Services."
    ),
    (
        "What academic support services are available to BIT students who are falling behind?",
        "the University offers Peer Assisted Study Sessions (PASS) facilitated by high-achieving senior students, Student Central for enrolment and study skills support, academic workshops on assignments, exam preparation, time management, and grammar, and consultation hours with course tutors and lecturers. Contact course staff or Student Support early rather than waiting."
    ),
    (
        "Are there student clubs relevant to IT or computing students at the University?",
        "Yes. the University has over 220 clubs and societies, including technology, coding, gaming, and entrepreneurship clubs relevant to IT and computing students. Explore available clubs at studentunion.example.edu/clubs-and-societies."
    ),
    (
        "What support does the University offer for BIT students with a disability or mental health condition?",
        "the University provides support, services, and facilities for students with a disability, illness, injury, or mental health condition. Contact a Diversity, Disability and Inclusion Adviser to arrange adjustments. More information is at my.example.edu/information-and-services/student-support/diversity-disability-inclusion."
    ),
    (
        "I need academic English support as a BIT student. What resources are available?",
        "the University offers a free an academic English program through the University College, delivered as workshops during semester. Academic English support covers a range of skills and topics. Contact Student Support and Wellbeing Services or check unicollege.edu.au for available options."
    ),
    (
        "What is the University Mentoring program, and how can BIT students access it?",
        "the University Mentoring connects students with peers, groups, or industry partners. It is particularly helpful for students new to university. As a BIT student you can be connected with mentors from IT or computing fields. More information is at my.example.edu/mentoring."
    ),
    (
        "Is there a Women in Computing program at the University for BIT students?",
        "Yes. the University's Women in Computing Program (WiC) encourages girls and women to pursue technology careers and fosters a sense of belonging for Computer Science and IT students. More information is available at dept.example.edu/wic."
    ),
    (
        "What library services are available to BIT students at the main campus?",
        "The main campus library provides academic information resources, welcoming study spaces, a world-class collection, and knowledgeable staff to assist with research and study questions. Online resources and study rooms are also available. More information is at library.example.edu."
    ),
    (
        "Does the University offer a student legal service?",
        "Yes. The University Student Union (the Student Union) can assist with legal or migration support. Contact studentunion.example.edu for details on available legal advice services for enrolled students."
    ),
    (
        "What is the University SafeZone app and when should I use it?",
        "The University SafeZone app connects students directly with the University security officers or emergency services from your mobile phone. It also provides a safety escort feature if you want someone to accompany you to public transport or your car. The Safety Bus also runs around the main campus from 6pm."
    ),
    (
        "What public transport options are available for getting to the main campus?",
        "the main campus is served by 10+ direct bus routes with buses arriving every 2 minutes at the University Lakes bus stop. There are 5 train stations within 4km, and a ferry service with 15-minute intervals. The campus is 7km from the city's city centre. Student concession transport cards reduce public transport costs."
    ),

    # ── I. CAREER OUTCOMES (Q129–Q143) ───────────────────────────────────────
    (
        "What percentage of BIT graduates from the University are in full-time employment?",
        "According to the University data, 82.5% of the University's Computer and Information Systems graduates are in full-time employment after graduation (based on QILT data). This is one of the stronger graduate employment rates for IT disciplines in Australia."
    ),
    (
        "What roles do BIT graduates from the University typically work in?",
        "BIT graduates pursue roles including software developer, experience designer, cyber security specialist, system architect, IT application specialist, product designer, and data governance manager. The broad curriculum across majors prepares graduates for diverse technology careers."
    ),
    (
        "What is the average salary for a software developer after graduating from the BIT at the University?",
        "Based on program information, software developers can earn approximately AUD $105,000 to $125,000 annually. Starting salaries for computing graduates from the University are reported at approximately $72,000 to $78,000, with specialist roles commanding higher remuneration."
    ),
    (
        "Are the University IT qualifications recognised internationally?",
        "Yes. the University's Engineering, Computer Science and IT faculty states that qualifications are internationally recognised and graduates can work anywhere in the world. the University is ranked among the global top 50 universities (QS 2025), supporting the international standing of its qualifications."
    ),
    (
        "Can I pursue further studies after completing the BIT?",
        "Yes. After the BIT you can apply for the Bachelor of IT (Honours) if you have a GPA of 5.0/7.0 or higher. You can also pursue postgraduate programs such as the Master of Information Technology (5581), Master of Data Science (5660), or Master of Cyber Security (5257) at the University."
    ),
    (
        "What is the Master of Information Technology at the University, and who is it designed for?",
        "The Master of IT (program code 5581) is a 2-year postgraduate program for students with a bachelor's degree in a field other than IT, computer science, software engineering, interaction design, or multimedia design. It covers Introduction to Software Engineering, Design Computing, Information Technology, and Relational Database Systems. The 2026 international fee is $58,056/year."
    ),
    (
        "What postgraduate IT options are available at the University for BIT graduates?",
        "After the BIT, you can pursue the BIT (Honours) or postgraduate programs including Master of IT (5581), Master of Interaction Design (5580), Master of Data Science (5660), Master of Cyber Security (5257), and Master of Computer Science (5522). Admission requirements vary by program."
    ),
    (
        "Does the BIT include any work-integrated learning or internship opportunities?",
        "the University supports work-integrated learning through enrichment programs, internships, global study, and industry events. Specific placement opportunities depend on your major and courses. Explore the University the career-development program at study.example.edu/enhance-your-employability for available options."
    ),
    (
        "What industries do BIT graduates from the University typically work in?",
        "BIT graduates work across technology, government, finance, healthcare, design, and retail. Dual degree graduates also enter industries related to their second degree - for example, health technology for Human Movement/IT graduates, or finance and banking for Commerce/IT graduates."
    ),
    (
        "Is the University's BIT ranked in world subject rankings?",
        "the University is ranked equal 40th globally (QS World University Rankings 2025). The BIT (Honours) program is ranked #1 in the state for computer science and information systems (QS World University Rankings 2026). the University's the School faculty has won 12 national teaching awards in the last 10 years."
    ),
    (
        "Does the BIT prepare students for careers in artificial intelligence or machine learning?",
        "The BIT curriculum includes AI content and Programming in the Large. For deeper specialisation, consider the Bachelor of Computer Science (AI major), or postgraduate options such as Master of Data Science or Master of Computer Science, which cover algorithms, AI, and machine learning."
    ),
    (
        "What career services does the University provide to help IT students find employment?",
        "the University provides career services through the career-development program, including global study, internships, career workshops, research projects, and industry events. The Faculty has industry connections through research partnerships and employer networks relevant to IT careers."
    ),
    (
        "Can I work in a part-time IT job while studying BIT on a student visa?",
        "Yes. On a student visa you can work up to 48 hours per fortnight during semester and full-time during holidays (from 1 July 2023). Part-time IT-related work such as casual software development, IT support, or data analysis is permitted within these hour limits. Ensure your employer is aware of your visa conditions."
    ),
    (
        "What career paths are typical for BIT/Arts dual degree graduates?",
        "BIT/Arts graduates can pursue cloud architecture, software engineering, data specialisation, UX design, and cybersecurity, while the Arts component opens pathways in communications, policy, education, and cross-cultural roles. This combination suits tech roles in government, media, international organisations, and education."
    ),
    (
        "What professional associations can BIT graduates join in Australia?",
        "The Australian Computer Society (ACS) is the primary professional body for ICT professionals in Australia. ACS membership is relevant for career development, networking, and certain visa pathways. The IEEE Computer Society is a major international option. You can also join specialist associations for cyber security, UX design, or data science depending on your major."
    ),

    # ── J. HONOURS PROGRAM (Q144–Q153) ───────────────────────────────────────
    (
        "What GPA do I need to apply for the Bachelor of Information Technology (Honours)?",
        "You need a minimum GPA of 5.0 on the University's 7-point scale to apply for the BIT (Honours). Bachelor's degrees completed more than 5 years before the intended commencement date are not eligible."
    ),
    (
        "What does the Bachelor of Information Technology (Honours) involve?",
        "The BIT (Honours) is a one-year full-time program (program code 9001) that includes a challenging research project developing communication skills, independence, creativity, and professional ethics. You can optionally integrate an industry placement into the research project."
    ),
    (
        "Can graduates from other universities apply for the University BIT (Honours)?",
        "Yes. Graduates from other universities can apply, but their bachelor's degree must be in IT or computer science and must include: algorithms and data structures, programming in C/C++/Java (or equivalent), discrete mathematics, web/mobile programming, information systems or databases, and human-computer interaction. The GPA requirement also applies."
    ),
    (
        "What are the application deadlines for the BIT (Honours)?",
        "For Semester 1 entry: 30 November (international) and 31 January (domestic). For Semester 2: 31 May (international) and 30 June (domestic). Submit applications via apply.example.edu using program code 9001."
    ),
    (
        "Is the Bachelor of Information Technology (Honours) available in Semester 2?",
        "Yes. The BIT (Honours) is available for both Semester 1 (23 February 2026) and Semester 2 (27 July 2026) intakes at the main campus."
    ),
    (
        "What is the annual tuition fee for the BIT (Honours) for international students?",
        "The international tuition fee for the BIT (Honours) is AUD $58,056 per year for 2026. Since the program is 1 year in duration, the total indicative cost is approximately AUD $58,056."
    ),
    (
        "What is the domestic tuition fee for the BIT (Honours)?",
        "The approximate annual domestic tuition fee for the BIT (Honours) is AUD $9,540 for 2026 (based on 16 units). Domestic students may be eligible for HECS-HELP if in a Commonwealth Supported Place."
    ),
    (
        "Can I do a Higher Degree by Research in IT at the University after the BIT?",
        "Yes. the University offers Doctor of Philosophy (PhD) and Master of Philosophy (MPhil) research degrees in IT and computing-related disciplines. The PhD typically requires a bachelor's with honours class IIA or equivalent research experience. International HDR students pay tuition in quarterly research quarters rather than semesters."
    ),
    (
        "Is the BIT (Honours) program ranked in any world rankings?",
        "Yes. The University BIT (Honours) program is ranked #1 in the state for computer science and information systems, according to the QS World University Rankings 2026."
    ),
    (
        "After completing BIT (Honours), what further study options are available?",
        "After BIT (Honours), you may be eligible for a Higher Degree by Research (PhD or MPhil) at the University or another university, or a postgraduate coursework degree such as Master of Data Science, Master of Cyber Security, or Master of Computer Science. Your research project may also open research roles in industry or academia."
    ),

    # ── K. APPLICATION PROCESS (Q154–Q163) ───────────────────────────────────
    (
        "How do domestic students apply for the BIT at the University?",
        "Domestic students apply through the state admissions centre (the state Tertiary Admissions Centre) at admissions.example.edu. The state admissions centre code for the Bachelor of Information Technology is 733001. Interstate students also apply through the state admissions centre."
    ),
    (
        "How do international students apply for the BIT at the University?",
        "International students apply directly to the University through the online portal at apply.example.edu. You need to create an account, complete your application with your qualifications and supporting documents, and pay the non-refundable AUD $150 application fee."
    ),
    (
        "Can I defer my the University BIT offer to a future semester if I cannot start in the offered semester?",
        "Contact the University International Admissions before the commencement date if you want to defer. If you have already received your Confirmation of Enrolment (CoE), deferral is only permitted under compassionate or compelling circumstances. Deferrals may affect your student visa, so seek advice before requesting one."
    ),
    (
        "Can I apply for multiple the University IT programs at the same time?",
        "International students can submit applications to multiple programs through apply.example.edu but can only hold one offer at a time. Domestic students list multiple preferences through the state admissions centre in order of preference. Seek advice from the state admissions centre or the University about how multiple applications are managed."
    ),
    (
        "What is a Genuine Student (GS) Assessment, and do I need to complete one?",
        "A Genuine Student Assessment is required by the Australian Government for international applicants to confirm they genuinely intend to study in Australia. You may be asked to complete a GS Assessment and provide a GS Agent Checklist (if using an agent) as part of your the University application. Failure to complete it may affect your application outcome."
    ),
    (
        "If I am still waiting for my final exam results, can I apply to the University for BIT?",
        "Yes. If you are waiting for transcripts or final results, you may receive a conditional offer. Submit your application with any available results and evidence of English language proficiency. Provide outstanding documents to your application portal as soon as you receive them."
    ),
    (
        "Where can I find a University-approved agent to help with my international application?",
        "the University works with approved agents around the world to assist international students with applications. Find a list of approved agents through the University International website. If you use an agent, you will also need to provide a GS Agent Checklist as part of your application."
    ),
    (
        "Can I submit my the University international application by post or in person instead of online?",
        "Yes. If you cannot apply online, submit in person or by mail to: Future Students and Admissions, Level 2, the Administration Building, the main campus, the University, the city, the state 0000, Australia. You can also email anon@example.edu to request an application form."
    ),
    (
        "How do I check the status of my the University BIT application?",
        "After submitting online, check the status through your account on the apply.example.edu portal. If you cannot access the portal, email anon@example.edu."
    ),
    (
        "I received an offer for the BIT from the University. What do I need to do to accept it?",
        "Follow the instructions in your offer letter, which typically include: accepting the offer through the apply.example.edu portal, paying any required tuition fee deposit, and arranging OSHC if you are an international student. Then apply for your student visa if required. Contact the University Admissions if you are unsure about any step."
    ),

    # ── L. ADDITIONAL TOPICS (Q164–Q200) ─────────────────────────────────────
    (
        "What is the difference between the Master of Information Technology and the Master of Computer Science at the University?",
        "The Master of IT (5581, 2 years) is for students with a background outside IT/CS and covers foundational IT topics. The Master of Computer Science (5522, 1.5 years) is for students who already hold a bachelor's in computer science or software engineering and covers advanced topics such as algorithms, AI, and machine learning. Both are at the main campus with 2026 international fees of $58,056/year."
    ),
    (
        "I completed a BIT and want to pursue a Master of Data Science at the University. What are the entry requirements?",
        "The Master of Data Science (5660) requires a bachelor's degree in a relevant discipline. As a BIT graduate, your eligibility will depend on whether your coursework includes sufficient mathematics, statistics, and computing content. Check the specific entry criteria on study.example.edu, as the MDS has two entry pathways depending on your undergraduate background."
    ),
    (
        "Can a BIT graduate apply for the Master of Cyber Security at the University?",
        "Yes. The Master of Cyber Security (5257) accepts applicants with a bachelor's degree in any discipline. BIT graduates are competitive applicants given their IT background. The program covers Fundamentals of Cyber Security, Information Security Essentials, Cyber Security Policy, Cyber Criminology, and Global Security, with majors in Cryptography, Cyber Criminology, Cyber Defence, and Leadership."
    ),
    (
        "What GPA do I need to apply for an IT-related PhD at the University?",
        "For PhD admission, you generally need a bachelor's degree with honours class IIA or better, or a research master's degree, or a coursework master's with a GPA equivalent to 5.65 on the University's 7-point scale. You also need a research proposal and must identify a potential supervisor. Contact anon@example.edu for guidance."
    ),
    (
        "I am an IT professional with five years of work experience but no honours degree. Can I apply for a PhD at the University?",
        "PhD entry is considered on a case-by-case basis. A bachelor's degree plus at least 2 years of documented relevant research experience including research publications may be considered equivalent to honours IIA. Contact the University Graduate School directly to discuss your situation before applying."
    ),
    (
        "Are there scholarships specifically available for IT students at the University?",
        "the University offers competitive scholarships covering academic excellence, sporting excellence, and study area. While IT-specific scholarships are not listed by name in the guide, check scholarships.example.edu for current opportunities relevant to IT students. Simply indicate scholarship consideration on your application - no separate form is required."
    ),
    (
        "What is the Concurrent Diploma in Languages, and can BIT students enrol in it?",
        "The Concurrent Diploma in Languages lets you learn a second language (9 available) alongside your BIT. It is 16 units and can be applied for after enrolling in a full-time the University bachelor's program. It adds language skills to your IT qualification without significantly extending your total study time."
    ),
    (
        "I am a BIT student interested in entrepreneurship. What opportunities does the University offer?",
        "the University actively supports entrepreneurship through free programs, clubs, and the University entrepreneurial ecosystem. Enrichment programs including internships, global study, research projects, and industry events contribute to entrepreneurial development. The University also supports entrepreneurs at every stage of their journey through industry networks."
    ),
    (
        "Does the University offer global study opportunities for BIT students?",
        "the University supports global study, internships, and volunteer opportunities through the career-development program. the University's 350,000 alumni across 190 countries and broad global partnerships create significant international opportunities. Enquire about specific exchange programs available to IT students at study.example.edu."
    ),
    (
        "What is the University's overall world ranking relevant to IT students?",
        "the University is ranked equal 40th in the world (QS World University Rankings 2025), 1st in Australia in the Nature Index Top Academic Institutions 2024, and 1st in Australia in the Australian Financial Review Best Universities Ranking 2023 and 2024. the University has 350,000 alumni from 190 countries."
    ),
    (
        "Is the BIT program accredited by the Australian Computer Society (ACS)?",
        "You should check the current accreditation status of the BIT on the program's page at study.example.edu or by contacting the School of the School, as accreditation status and conditions can change. ACS accreditation affects eligibility for professional ACS membership after graduation."
    ),
    (
        "Can I transfer from the BIT at the University to the Bachelor of Computer Science during my studies?",
        "A transfer may be possible, depending on which courses you have completed and whether they satisfy BCS entry and credit requirements. Contact the School of the School or the Faculty for formal academic advice before attempting a transfer, as the process is not automatic."
    ),
    (
        "What is the University grading scale, and what does a grade of 4 mean?",
        "the University uses a 7-point grading scale where 4 is a pass and 7 is the highest grade (equivalent to High Distinction). Grades below 4 are failing grades. For postgraduate admission, the University typically requires a minimum GPA of 5.0/7.0. Failing grades are included in GPA calculations."
    ),
    (
        "Does the BIT include courses on social media and mobile computing?",
        "Yes. According to the University International Guide 2026, Social and Mobile Computing is listed as one of the key areas of study in the BIT curriculum (program code 9002), alongside Human-Computer Interaction, Programming in the Large, and Artificial Intelligence."
    ),
    (
        "What are the main differences between the BIT and the Bachelor of Computer Science at the University?",
        "The BIT (code 9002, 3 years) focuses on Human-Computer Interaction, Social and Mobile Computing, AI, and Programming in the Large, with majors in Software Design, Software Information Systems, and User Experience Design. The Bachelor of Computer Science (code 2451, 3 years) covers Software Engineering, Information Systems, Discrete Mathematics, and Computer Systems, with majors in Cyber Security, Data Science, Artificial Intelligence, Programming Languages, and Scientific Computing."
    ),
    (
        "Does the city have a strong tech industry for IT graduates from the University?",
        "Yes. the city is one of Australia's fastest-growing capital cities (ABS 2023) and one of the top 30 best student cities in the world (QS 2025). It is a thriving centre in Asia-Pacific for technology, investment, and major events. With 82.5% of the University IT graduates in full-time employment, the local and regional job market is strong."
    ),
    (
        "Is cyber security covered in the BIT curriculum?",
        "Cyber security career pathways are listed for the School graduates. While the BIT does not have a dedicated Cyber Security major (unlike the Bachelor of Computer Science), relevant content and electives may be available. For deeper specialisation, consider the BCS Cyber Security major or the postgraduate Master of Cyber Security (5257)."
    ),
    (
        "What is the University Student Union and what does it offer BIT students?",
        "The University Student Union (the Student Union) is the student representative body offering legal and migration support, clubs and societies, events, and student advocacy. BIT students can join IT and computing clubs through the Student Union and access legal advice at studentunion.example.edu."
    ),
    (
        "What are the key semester dates for 2026 at the University?",
        "Key 2026 dates: Semester 1 starts 23 February; Semester 1 census date 31 March; Semester 1 exam period 6–20 June; Semester 2 starts 27 July; Semester 2 census date 31 August; Semester 2 exam period 7–21 November. International application closing dates: 30 November (Semester 1) and 31 May (Semester 2)."
    ),
    (
        "What is the census date and why does it matter for BIT students?",
        "The census date is the deadline for finalising enrolment. For domestic students, withdrawing before the census date means you will not incur a HECS-HELP debt for that course. For international students, the census date also affects fee refunds and visa considerations. For 2026: Semester 1 census date is 31 March; Semester 2 census date is 31 August."
    ),
    (
        "Can a BIT student add a second major within the 3-year program?",
        "The BIT is a 3-year single-major program and its standard unit count typically does not accommodate a formal second major. If you want to broaden your studies, consider a dual degree, a Concurrent Diploma in Languages, or strategic use of elective units. Seek academic advice to explore your options."
    ),
    (
        "What is the PASS program at the University, and how can it help BIT students?",
        "Peer Assisted Study Sessions (PASS) are facilitated by second- and third-year students who achieved top grades in specific courses. They provide advice on course content and study habits. PASS is particularly useful for first-year BIT courses like CORE1001 and INFO1200. Ask your faculty for details on PASS sessions for your courses."
    ),
    (
        "Can I study the BIT entirely online?",
        "No. The BIT is a campus-based program delivered at the main campus. Some course components may have online elements through the University's digitally enhanced learning environment, but the program is not fully online. International students on a student visa must also study on-campus."
    ),
    (
        "I am applying for BIT from Vietnam with a Bang Tu Tai diploma. Am I eligible?",
        "Vietnamese students who have completed the Bang Tu Tai or Bang Tot Nghiep Pho Thong Trung Hoc (Upper Secondary School Graduation Diploma) will be considered for most programs, provided the required Year 12 grade is obtained. Contact the University International Admissions at anon@example.edu to confirm your eligibility."
    ),
    (
        "Can work experience be counted as prior learning credit when applying for the BIT as an international student?",
        "Credit for work experience alone is not typically granted for undergraduate programs at the University. Credit is assessed based on formal prior learning that is equivalent in level, standard, and content to the University courses. Check the credit eligibility database at creditprecedents.app.example.edu to see whether your prior study qualifies."
    ),
    (
        "What is the difference between BIT program codes 9002 and 9004?",
        "Both codes appear in the University documentation as the Bachelor of Information Technology. Code 9002 (CRICOS CRIC001) appears in the University International Guide 2026, while 9004 appears on the study.example.edu program page for 2026 entry. These may reflect a program update. Refer to study.example.edu for the current program code and confirm with admissions if uncertain."
    ),
    (
        "What are the Semester 1 and Semester 2 start dates at the University for 2026?",
        "Semester 1, 2026 commences 23 February 2026, with O-Week from 16–20 February. Semester 2, 2026 commences 27 July 2026, with Mid-year Orientation Week from 20–24 July. These dates are important for planning your arrival and accommodation as a new student."
    ),
    (
        "How long will the BIT take if I study part-time as a domestic student?",
        "Part-time study is defined as less than 75% of the standard full-time load (fewer than 6 units per semester). Studying part-time will extend the BIT beyond the standard 3 years. Plan your course sequence carefully to meet all prerequisites in the correct order. Domestic students can study part-time without visa restrictions."
    ),
    (
        "What majors are available in the Bachelor of Computer Science at the University (for comparison with BIT)?",
        "The Bachelor of Computer Science (2451) offers majors in Cyber Security, Data Science, Artificial Intelligence, Programming Languages, and Scientific Computing. This contrasts with the BIT, which offers Software Information Systems, User Experience Design, and Software Design/Solution Architecture. Choose between them based on whether you prefer IT business/design applications or core computing and algorithms."
    ),
    (
        "Does the University provide tutoring support for BIT courses such as CORE1001 or DSGN1400?",
        "the University offers Peer Assisted Study Sessions (PASS) facilitated by high-achieving peers, consultation hours with lecturers and tutors, and online course forums. Attending weekly practicals and labs also provides hands-on guidance. Contact course staff early if you are struggling rather than waiting until assessment time."
    ),
    (
        "Is the BIT degree useful for applying for a post-study work visa in Australia after graduation?",
        "Completing a 3-year BIT at the University may make you eligible for a Temporary Graduate Visa (subclass 485) after graduation, allowing you to work temporarily in Australia. The duration depends on qualification level and study location. Check the most current visa requirements with the Australian Department of Home Affairs at homeaffairs.gov.au, as migration law changes frequently."
    ),
    (
        "I completed my BIT at the University and want to apply for the BIT (Honours). Do I need to apply again?",
        "Yes. The BIT (Honours) is a separate program and requires a new application. Apply online at apply.example.edu using program code 9001. For domestic students, the Semester 1 deadline is 31 January; for international students, 30 November of the previous year. You must meet the minimum GPA of 5.0/7.0 from your BIT."
    ),
    (
        "What is DSGN3801 and why is it important in the BIT degree?",
        "DSGN3801 (Design & Computation Studio 2 / Projects in Computing) is a capstone project course in the BIT. It requires students to work in multidisciplinary teams to design, prototype, and evaluate a computing product for a real client. It integrates skills from across the degree and is typically taken in the final year. It is a core requirement for BIT students and is not offered in Summer Semester."
    ),
    (
        "Can I do a study abroad or exchange semester as a BIT student at the University?",
        "Yes. the University encourages students to participate in exchange programs through the University Abroad office. As a BIT student you can apply for semester-long exchanges at partner universities worldwide. Credit for courses taken abroad is assessed by the BIT program coordinator. Contact the University Abroad (studyabroad.example.edu) early - applications typically open 9–12 months before the planned exchange semester."
    ),
    (
        "How does academic integrity work at the University and what counts as misconduct in BIT courses?",
        "the University's Student Integrity and Misconduct Policy covers all forms of academic dishonesty, including plagiarism, contract cheating, collusion, and misrepresentation. In BIT courses this includes submitting code written by others or using AI tools without declared authorisation. Each course statement specifies what is and is not permitted. Penalties range from a mark of zero to expulsion. Submit all work through the correct channels and use the University's academic integrity resources at academicintegrity.example.edu."
    ),
    (
        "What are the prerequisites for INFO1200 Introduction to Information Systems?",
        "INFO1200 has no listed prerequisites at the University and is designed as a first-year introductory course. It covers database concepts, entity-relationship modelling, SQL, and data management. It is a required course for most BIT study plans and should ideally be completed in your first year. Completing INFO1200 is a prerequisite for INFO2200 (Relational Database Systems) in second year."
    ),
    (
        "Is there a difference in tuition fees between BIT majors or dual degree programs?",
        "For international students, tuition fees for the BIT and all dual degrees with BIT are charged per unit and can differ between programs depending on the faculty administering each component. For example, engineering courses in the BIT/Engineering dual degree may attract higher engineering unit fees. Domestic students pay HECS-HELP student contribution amounts that vary by course band. Always check the specific fee schedule for your program at fees.example.edu before enrolling."
    ),
]

assert len(QA_PAIRS) == 200, f"Expected 200 Q&A pairs, got {len(QA_PAIRS)}"

def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, heading in enumerate(["Question", "Answer"], start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Data rows
    for row_idx, (q, a) in enumerate(QA_PAIRS, start=2):
        for col, val in enumerate([q, a], start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Alternate row shading
        if row_idx % 2 == 0:
            shade = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            for col in range(1, 3):
                ws.cell(row=row_idx, column=col).fill = shade

    # Column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 90

    # Freeze header
    ws.freeze_panes = "A2"

    return wb


if __name__ == "__main__":
    import os
    out_path = os.path.join(os.path.dirname(__file__), "generated-qa-200.xlsx")
    wb = create_workbook()
    wb.save(out_path)
    print(f"Saved {len(QA_PAIRS)} Q&A pairs to: {out_path}")
