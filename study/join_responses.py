#!/usr/bin/env python
"""
study/join_responses.py
=======================
Joins returned Excel questionnaire files with study/answer_key.csv to
produce a tidy long table ready for analysis.

Usage:
    python study/join_responses.py \\
        --responses-dir study/returned \\
        [--key  study/answer_key.csv]  \\
        [--out  study/responses_tidy.csv]

Each returned workbook is one rater's filled copy of study/questionnaire.xlsx.

Output columns (one row per rater x item_id x dimension):
    rater, item_id, dimension, score,
    model, state, variant, condition, source_file,
    composite_score_geval, factual_accuracy_geval, relevance_geval,
    conciseness_geval, no_hallucination_geval,
    question_slot, question_idx, display_letter

How the join works:
    Each returned workbook contains a very-hidden '_MAP' sheet with columns
    (tag, sheet, cell). This script reads that map, opens the (sheet, cell)
    pair to get the rater's score, parses the item_id and dimension code from
    the tag using a regex, then joins condition metadata and G-Eval scores
    from answer_key.csv via item_id.

Rater identification:
    The rater's email is read from cell B14 of the 'Start here' sheet
    (where they were asked to enter it). Fallback order: B13 (name), file stem.

SEED = 42  (used in build_study.py; recorded here for reference only —
            this script performs no randomisation).
"""
import argparse
import pathlib
import re
import sys

import openpyxl
import pandas as pd

# Matches "[Q01_A · FACT]", "[Q10_E · NOHAL]", etc.
TAG_RE = re.compile(
    r"\[([A-Z0-9_]+)\s*[·•\-]\s*([A-Z]+)\]",
    re.UNICODE,
)

# Must match COVER_NAME_CELL / COVER_EMAIL_CELL in build_study.py
COVER_SHEET  = "Start here"
MAP_SHEET    = "_MAP"
NAME_CELL    = "B13"
EMAIL_CELL   = "B14"

TIDY_COLS = [
    "rater", "item_id", "dimension", "score",
    "model", "state", "variant", "condition", "source_file",
    "composite_score_geval", "factual_accuracy_geval", "relevance_geval",
    "conciseness_geval", "no_hallucination_geval",
    "question_slot", "question_idx", "display_letter",
]


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        description="Join returned Excel questionnaires with answer key to a tidy CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--responses-dir", required=True, metavar="DIR",
        help="Folder containing the returned .xlsx questionnaire files.",
    )
    p.add_argument(
        "--key", default="study/answer_key.csv", metavar="CSV",
        help="Path to the private answer key (default: study/answer_key.csv).",
    )
    p.add_argument(
        "--out", default="study/responses_tidy.csv", metavar="CSV",
        help="Output path for the tidy long CSV (default: study/responses_tidy.csv).",
    )
    return p.parse_args(argv)


def get_rater_id(wb, file_path):
    """Return email (B14), then name (B13), then file stem as rater identifier."""
    try:
        ws = wb[COVER_SHEET]
        email = ws[EMAIL_CELL].value
        if email and str(email).strip():
            return str(email).strip()
        name = ws[NAME_CELL].value
        if name and str(name).strip():
            return str(name).strip()
    except Exception:
        pass
    return pathlib.Path(file_path).stem


def read_workbook(wb, rater, key_lookup):
    """Return tidy rows from a single filled workbook."""
    rows = []

    try:
        ws_map = wb[MAP_SHEET]
    except KeyError:
        print(
            f"WARNING: workbook for '{rater}' has no '{MAP_SHEET}' sheet — skipping.",
            file=sys.stderr,
        )
        return rows

    for map_row in ws_map.iter_rows(min_row=2, values_only=True):
        if not map_row or map_row[0] is None:
            continue
        tag, sheet_name, cell_addr = map_row[0], map_row[1], map_row[2]
        if not tag or not sheet_name or not cell_addr:
            continue

        m = TAG_RE.search(str(tag))
        if not m:
            continue
        item_id, dim_code = m.group(1), m.group(2)

        try:
            score_raw = wb[sheet_name][cell_addr].value
            score = int(score_raw) if score_raw is not None else None
        except (KeyError, ValueError, TypeError):
            score = None

        meta = key_lookup.get(item_id, {})
        rows.append({
            "rater":                  rater,
            "item_id":                item_id,
            "dimension":              dim_code,
            "score":                  score,
            "model":                  meta.get("model", ""),
            "state":                  meta.get("state", ""),
            "variant":                meta.get("variant", ""),
            "condition":              meta.get("condition", ""),
            "source_file":            meta.get("source_file", ""),
            "composite_score_geval":  meta.get("composite_score", ""),
            "factual_accuracy_geval": meta.get("factual_accuracy", ""),
            "relevance_geval":        meta.get("relevance", ""),
            "conciseness_geval":      meta.get("conciseness", ""),
            "no_hallucination_geval": meta.get("no_hallucination", ""),
            "question_slot":          meta.get("question_slot", ""),
            "question_idx":           meta.get("question_idx", ""),
            "display_letter":         meta.get("display_letter", ""),
        })

    return rows


def main(argv=None):
    args = parse_args(argv)

    # Load answer key
    try:
        key_df = pd.read_csv(args.key)
    except FileNotFoundError:
        print(f"ERROR: answer key not found at '{args.key}'", file=sys.stderr)
        sys.exit(1)
    key_lookup = key_df.set_index("item_id").to_dict(orient="index")

    # Find returned xlsx files
    responses_dir = pathlib.Path(args.responses_dir)
    if not responses_dir.is_dir():
        print(f"ERROR: '{responses_dir}' is not a directory.", file=sys.stderr)
        sys.exit(1)
    xlsx_files = sorted(responses_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"ERROR: No .xlsx files found in '{responses_dir}'.", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(xlsx_files)} .xlsx file(s) in '{responses_dir}'.")

    tidy_rows = []
    for f in xlsx_files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            print(f"WARNING: Could not open '{f.name}': {e}", file=sys.stderr)
            continue
        rater = get_rater_id(wb, f)
        rows = read_workbook(wb, rater, key_lookup)
        n_filled = sum(1 for r in rows if r["score"] is not None)
        print(f"  {f.name}: rater='{rater}', {len(rows)} items mapped, {n_filled} filled.")
        tidy_rows.extend(rows)

    if not tidy_rows:
        print(
            "ERROR: No data collected. Check that files were built from questionnaire.xlsx "
            "and contain a '_MAP' sheet.",
            file=sys.stderr,
        )
        sys.exit(1)

    tidy_df = pd.DataFrame(tidy_rows, columns=TIDY_COLS)
    tidy_df.to_csv(args.out, index=False)

    n_raters  = tidy_df["rater"].nunique()
    n_items   = len(tidy_df)
    n_missing = int(tidy_df["score"].isna().sum())

    print()
    print(f"Output: {n_items} rows x {len(tidy_df.columns)} cols  ->  {args.out}")
    print(f"  {n_raters} unique rater(s),  {n_items} item scores,  {n_missing} missing.")

    if n_missing:
        print(
            f"WARNING: {n_missing} score(s) are blank — "
            "consider following up with those raters.",
            file=sys.stderr,
        )

    # Warn about any item_ids not found in the answer key
    unknown = tidy_df[tidy_df["model"] == ""]["item_id"].unique()
    if len(unknown):
        print(
            f"WARNING: {len(unknown)} item_id(s) not found in answer key: {sorted(unknown)}",
            file=sys.stderr,
        )

    print()
    print(
        "Column guide:\n"
        "  rater            -- rater email or identifier\n"
        "  item_id          -- opaque answer ID (Q01_A ... Q10_E)\n"
        "  dimension        -- FACT | REL | CONC | NOHAL\n"
        "  score            -- human rating 0-5 (integer, or blank if not filled)\n"
        "  model/state/variant/condition -- condition metadata from answer_key.csv\n"
        "  composite_score_geval etc.    -- original G-Eval scores for comparison\n"
        "  question_slot/question_idx/display_letter -- form context"
    )


if __name__ == "__main__":
    main()
