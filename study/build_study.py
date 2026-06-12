#!/usr/bin/env python
"""
study/build_study.py
====================
Generates two study artefacts:

  1. study/answer_key.csv  — PRIVATE. Never share with raters.
     Maps every opaque item_id to its model / state / variant / condition
     and records the display-letter assignment.

  2. study/questionnaire.xlsx  — The rater-facing questionnaire workbook.
     10 question sheets (Q01-Q10), each showing the student question,
     reference answer, and a 5-answer x 4-metric grid of yellow score cells.
     Distribute one copy per rater; they fill it and return it by email.

Deterministic: every randomisation step uses SEED = 42.

Usage (from repo root):
    python study/build_study.py
"""
import csv
import json
import pathlib

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

# ── Constants ─────────────────────────────────────────────────────────────────
SEED = 42
REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_JSONL = REPO_ROOT / "data" / "test.jsonl"
OUT_DIR = pathlib.Path(__file__).resolve().parent

MODELS = [
    "gemma3-12b-grpo",
    "mistral-nemo-12b-grpo",
    "qwen3-14B-grpo",
]
STATES = ["base", "finetuned"]
VARIANTS = [
    "plain",
    "sysprompt",
    "fewshot",
    "sysprompt_fewshot",
    "rag",
    "sysprompt_rag",
    "rag_fewshot",
    "sysprompt_rag_fewshot",
]

DIMENSIONS = [
    ("FACT",  "Factual accuracy",
     "Does the answer contain only accurate, verifiable facts about the UQ BIT program? "
     "(0 = completely inaccurate or misleading, 5 = fully accurate)"),
    ("REL",   "Relevance",
     "Does the answer directly address what the student asked? "
     "(0 = completely off-topic, 5 = perfectly on-point)"),
    ("CONC",  "Conciseness",
     "Is the answer appropriately brief without omitting essential information? "
     "(0 = very padded or repetitive, 5 = ideally concise)"),
    ("NOHAL", "No hallucination",
     "Does the answer avoid fabricated details, invented course codes, or made-up policies? "
     "(0 = heavily hallucinated, 5 = no hallucinations detected)"),
]

ATTN_SLOTS = {}  # no attention checks

MAX_ANSWER_CHARS = 3000

ANSWER_KEY_COLS = [
    "item_id",
    "question_slot",
    "question_idx",
    "display_letter",
    "model",
    "state",
    "variant",
    "condition",
    "source_file",
    "composite_score",
    "factual_accuracy",
    "relevance",
    "conciseness",
    "no_hallucination",
]

# Cell addresses used in the 'Start here' cover sheet (must match generate_workbook).
COVER_NAME_CELL  = "B13"
COVER_EMAIL_CELL = "B14"


# ── Load test.jsonl ───────────────────────────────────────────────────────────
def load_test_records():
    recs = []
    with DATA_JSONL.open(encoding="utf-8") as fh:
        for line in fh:
            recs.append(json.loads(line))
    return recs


# ── Load all 48 result CSVs ───────────────────────────────────────────────────
def load_all_results():
    frames = []
    for model in MODELS:
        for state in STATES:
            for variant in VARIANTS:
                rel_path = f"fine-tuning/{model}/results/eval_{state}_{variant}.csv"
                path = REPO_ROOT / rel_path
                df = pd.read_csv(path)
                df = df[df["idx"].astype(str).str.strip().str.isdigit()].copy()
                df["idx"] = df["idx"].astype(int)
                df["model"] = model
                df["state"] = state
                df["variant"] = variant
                df["source_file"] = rel_path
                frames.append(df)
    return pd.concat(frames, ignore_index=True)


# ── Sample 10 questions (seed 42) ─────────────────────────────────────────────
def sample_questions(all_idx, rng):
    sampled = sorted(rng.choice(sorted(all_idx), size=10, replace=False).tolist())
    return sampled


# ── Pick 5 conditions spanning the composite range ────────────────────────────
def pick_five_conditions(sub_df):
    sub = sub_df.sort_values("composite_score").reset_index(drop=True)
    cmin = sub["composite_score"].min()
    cmax = sub["composite_score"].max()
    targets = np.linspace(cmin, cmax, 5)
    chosen = []
    used_indices = set()
    for t in targets:
        cand = sub[~sub.index.isin(used_indices)].copy()
        cand["_dist"] = (cand["composite_score"] - t).abs()
        best_idx = cand["_dist"].idxmin()
        used_indices.add(best_idx)
        chosen.append(sub.loc[best_idx])
    return chosen


# ── Assign display letters A-E (seed 42, independent per question) ────────────
def make_letter_assignments(n_questions, rng):
    return [[str(l) for l in rng.permutation(["A", "B", "C", "D", "E"])]
            for _ in range(n_questions)]


# ── Build answer-key rows and per-question study data ─────────────────────────
def build_study(test_records, long_df, sampled_idx, letter_assignments):
    answer_key_rows = []
    study_questions = []

    for slot_0, test_idx in enumerate(sampled_idx):
        slot = slot_0 + 1

        question_text = test_records[test_idx]["messages"][1]["content"]
        reference_text = test_records[test_idx]["messages"][2]["content"]

        sub = long_df[long_df["idx"] == test_idx].copy()
        five = pick_five_conditions(sub)
        letters = letter_assignments[slot_0]

        cond_letter_pairs = list(zip(five, letters))

        answers_for_form = []
        for row, letter in cond_letter_pairs:
            item_id = f"Q{slot:02d}_{letter}"
            answer_text = str(row["generated_response"])
            if len(answer_text) > MAX_ANSWER_CHARS:
                answer_text = answer_text[:MAX_ANSWER_CHARS] + " ...[truncated]"

            answer_key_rows.append({
                "item_id":          item_id,
                "question_slot":    slot,
                "question_idx":     int(test_idx),
                "display_letter":   letter,
                "model":            row["model"],
                "state":            row["state"],
                "variant":          row["variant"],
                "condition":        f"{row['state']}_{row['variant']}",
                "source_file":      row["source_file"],
                "composite_score":  round(float(row["composite_score"]), 6),
                "factual_accuracy": round(float(row["factual_accuracy"]), 4),
                "relevance":        round(float(row["relevance"]), 4),
                "conciseness":      round(float(row["conciseness"]), 4),
                "no_hallucination": round(float(row["no_hallucination"]), 4),
            })
            answers_for_form.append({
                "letter":  letter,
                "item_id": item_id,
                "text":    answer_text,
            })

        answers_for_form.sort(key=lambda a: a["letter"])

        study_questions.append({
            "slot":      slot,
            "test_idx":  int(test_idx),
            "question":  question_text,
            "reference": reference_text,
            "answers":   answers_for_form,
        })

    answer_key_rows.sort(
        key=lambda r: (int(r["question_slot"]), str(r["display_letter"]))
    )

    return answer_key_rows, study_questions


# ── Write answer_key.csv ──────────────────────────────────────────────────────
def write_answer_key(rows, path):
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=ANSWER_KEY_COLS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[OK] answer_key.csv written  ({len(rows)} rows = 50 answers)")
    print(f"  -> {path}")


# ── Generate questionnaire.xlsx ───────────────────────────────────────────────
def generate_workbook(study_questions, path):
    # Styles
    YELLOW   = PatternFill("solid", fgColor="FFF2CC")
    HFILL    = PatternFill("solid", fgColor="1F4E79")  # dark-blue header
    QFILL    = PatternFill("solid", fgColor="DEEAF1")  # light-blue Q/ref rows
    LFILL    = PatternFill("solid", fgColor="E2EFDA")  # light-green answer-letter cell

    BOLD_WH  = Font(bold=True, color="FFFFFF", size=11)
    BOLD     = Font(bold=True, size=11)
    TITLE    = Font(bold=True, size=14)

    WRAP_TOP = Alignment(wrap_text=True, vertical="top")
    CTR      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    UNLOCKED = Protection(locked=False)

    # Dimension code -> column index (C=3, D=4, E=5, F=6)
    DIM_COLS = {"FACT": 3, "REL": 4, "CONC": 5, "NOHAL": 6}

    wb = openpyxl.Workbook()

    # ── Cover sheet ('Start here') ────────────────────────────────────────────
    ws_c = wb.active
    ws_c.title = "Start here"
    ws_c.column_dimensions["A"].width = 22
    ws_c.column_dimensions["B"].width = 78

    # Row 1: Title
    ws_c.merge_cells("A1:B1")
    ws_c["A1"] = "UQ BIT Information Assistant — Expert Quality Evaluation"
    ws_c["A1"].font = TITLE
    ws_c["A1"].alignment = CTR
    ws_c.row_dimensions[1].height = 35

    # Row 3: Purpose
    ws_c["A3"] = "PURPOSE"
    ws_c["A3"].font = BOLD
    ws_c["B3"] = (
        "This questionnaire is part of a research study at the University of Queensland "
        "evaluating the quality of AI-generated responses to student queries about the "
        "Bachelor of Information Technology (BIT) program."
    )
    ws_c["B3"].alignment = WRAP_TOP
    ws_c.row_dimensions[3].height = 48

    # Row 5: What you will do
    ws_c["A5"] = "WHAT YOU WILL DO"
    ws_c["A5"].font = BOLD
    ws_c["B5"] = (
        "You will review 10 student questions (one per sheet Q01–Q10). "
        "For each question you will see the student question, a reference answer "
        "from official UQ sources, and five AI-generated answers labelled A to E "
        "(in random order).\n\n"
        "For each answer, rate it on four dimensions using the 0–5 scale "
        "in the yellow cells:\n"
        "  •  Factual accuracy — Are the stated facts correct?\n"
        "  •  Relevance — Does it address the student’s question?\n"
        "  •  Conciseness — Is it appropriately brief?\n"
        "  •  No hallucination — Does it avoid fabricated details?"
    )
    ws_c["B5"].alignment = WRAP_TOP
    ws_c.row_dimensions[5].height = 115

    # Row 7: Estimated time
    ws_c["A7"] = "ESTIMATED TIME"
    ws_c["A7"].font = BOLD
    ws_c["B7"] = "40–60 minutes."

    # Row 9: Consent
    ws_c["A9"] = "CONSENT"
    ws_c["A9"].font = BOLD
    ws_c["B9"] = (
        "By filling in and returning this file you confirm that you are participating "
        "voluntarily and you consent to your anonymised responses being used in this "
        "research study conducted at the University of Queensland."
    )
    ws_c["B9"].alignment = WRAP_TOP
    ws_c.row_dimensions[9].height = 55

    # Row 11: Separator
    ws_c.merge_cells("A11:B11")
    ws_c["A11"] = "─" * 70
    ws_c.row_dimensions[11].height = 8

    # Rows 12-14: Rater info
    ws_c["A12"] = "RATER INFORMATION"
    ws_c["A12"].font = BOLD
    ws_c["A13"] = "Your name:"
    ws_c["A13"].font = BOLD
    ws_c[COVER_NAME_CELL].fill = YELLOW
    ws_c[COVER_NAME_CELL].protection = UNLOCKED
    ws_c["A14"] = "Your email:"
    ws_c["A14"].font = BOLD
    ws_c[COVER_EMAIL_CELL].fill = YELLOW
    ws_c[COVER_EMAIL_CELL].protection = UNLOCKED

    # Row 16: Scoring guide
    ws_c["A16"] = "SCORING"
    ws_c["A16"].font = BOLD
    ws_c["B16"] = (
        "Enter a whole number 0–5 in every yellow cell on sheets Q01–Q10.\n"
        "0 = lowest quality,  5 = highest quality.\n"
        "Excel will warn you if a value is outside 0–5."
    )
    ws_c["B16"].alignment = WRAP_TOP
    ws_c.row_dimensions[16].height = 50

    # Rows 18+: Dimension guide
    ws_c["A18"] = "DIMENSIONS"
    ws_c["A18"].font = BOLD
    for i, (code, label, help_text) in enumerate(DIMENSIONS):
        r = 19 + i
        ws_c.cell(row=r, column=1, value=f"{label} ({code})").font = BOLD
        ws_c.cell(row=r, column=2, value=help_text).alignment = WRAP_TOP
        ws_c.row_dimensions[r].height = 30

    ws_c.protection.sheet = True

    # ── Question sheets (Q01-Q10) ─────────────────────────────────────────────
    map_data = []  # rows for the _MAP sheet: [tag, sheet_name, cell_addr]

    for q in study_questions:
        slot = q["slot"]
        sheet_name = f"Q{slot:02d}"
        ws = wb.create_sheet(title=sheet_name)

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 78
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18

        # Row 1: Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Question {slot} of {len(study_questions)}"
        ws["A1"].font = TITLE
        ws["A1"].alignment = CTR
        ws.row_dimensions[1].height = 28

        # Row 2: Student question
        ws["A2"] = "STUDENT QUESTION:"
        ws["A2"].font = BOLD
        ws["A2"].fill = QFILL
        ws["A2"].alignment = WRAP_TOP
        ws.merge_cells("B2:F2")
        ws["B2"] = q["question"]
        ws["B2"].fill = QFILL
        ws["B2"].alignment = WRAP_TOP
        q_lines = max(3, len(q["question"]) // 76 + 1)
        ws.row_dimensions[2].height = max(45, min(150, q_lines * 14))

        # Row 3: Reference answer
        ws["A3"] = "REFERENCE ANSWER\n(official UQ sources):"
        ws["A3"].font = BOLD
        ws["A3"].fill = QFILL
        ws["A3"].alignment = WRAP_TOP
        ws.merge_cells("B3:F3")
        ws["B3"] = q["reference"]
        ws["B3"].fill = QFILL
        ws["B3"].alignment = WRAP_TOP
        ref_lines = max(3, len(q["reference"]) // 76 + 1)
        ws.row_dimensions[3].height = max(45, min(250, ref_lines * 14))

        # Row 4: Blank separator
        ws.row_dimensions[4].height = 6

        # Row 5: Grid column headers
        headers = [
            "Answer",
            "Answer text",
            f"Factual accuracy\n(0–5)",
            f"Relevance\n(0–5)",
            f"Conciseness\n(0–5)",
            f"No hallucination\n(0–5)",
        ]
        for col_i, h in enumerate(headers, start=1):
            c = ws.cell(row=5, column=col_i, value=h)
            c.font = BOLD_WH
            c.fill = HFILL
            c.alignment = CTR
        ws.row_dimensions[5].height = 35

        # Freeze rows 1-5 so headers stay visible when scrolling
        ws.freeze_panes = "A6"

        # Data validation for all score cells on this sheet
        dv = DataValidation(
            type="whole", operator="between",
            formula1="0", formula2="5",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="warning",
            errorTitle="Score out of range",
            error="Please enter a whole number between 0 and 5.",
            showInputMessage=True,
            promptTitle="Score",
            prompt="Enter 0 (lowest) to 5 (highest).",
        )
        ws.add_data_validation(dv)

        # Rows 6-10: One row per answer A-E
        for ans_i, ans in enumerate(q["answers"]):
            data_row = 6 + ans_i

            # Col A: answer letter
            c_letter = ws.cell(row=data_row, column=1, value=ans["letter"])
            c_letter.font = BOLD
            c_letter.fill = LFILL
            c_letter.alignment = CTR

            # Col B: answer text (wrapped)
            c_text = ws.cell(row=data_row, column=2, value=ans["text"])
            c_text.alignment = WRAP_TOP

            text_lines = max(4, len(ans["text"]) // 76 + 1)
            ws.row_dimensions[data_row].height = max(60, min(400, text_lines * 14))

            # Cols C-F: yellow score input cells
            for code, _label, _help in DIMENSIONS:
                col_i = DIM_COLS[code]
                c_score = ws.cell(row=data_row, column=col_i)
                c_score.fill = YELLOW
                c_score.protection = UNLOCKED
                c_score.alignment = CTR
                dv.add(c_score)

                cell_addr = f"{get_column_letter(col_i)}{data_row}"
                map_data.append([
                    f"[{ans['item_id']} · {code}]",
                    sheet_name,
                    cell_addr,
                ])

        ws.protection.sheet = True

    # ── _MAP sheet (very hidden — join reference, not visible to raters) ───────
    ws_map = wb.create_sheet(title="_MAP")
    ws_map.append(["tag", "sheet", "cell"])
    for row in map_data:
        ws_map.append(row)
    ws_map.sheet_state = "veryHidden"

    wb.save(path)
    print(f"[OK] questionnaire.xlsx written  (200 score cells, 10 question sheets)")
    print(f"  -> {path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"[build_study.py]  SEED = {SEED}")
    print()

    rng = np.random.RandomState(SEED)

    print("Loading data/test.jsonl ...")
    test_records = load_test_records()
    print(f"  {len(test_records)} test records loaded.")

    print("Loading 48 result CSVs ...")
    long_df = load_all_results()
    print(f"  {len(long_df)} (question x condition) rows loaded.")

    print()
    all_idx = sorted(long_df["idx"].unique())

    sampled_idx = sample_questions(all_idx, rng)
    print(f"Sampled question indices (seed {SEED}): {sampled_idx}")

    letter_assignments = make_letter_assignments(len(sampled_idx), rng)
    for slot_0, (idx, letters) in enumerate(zip(sampled_idx, letter_assignments)):
        q_text = test_records[idx]["messages"][1]["content"][:60]
        print(f"  Q{slot_0+1:02d}  idx={idx:2d}  letters={letters}  \"{q_text}...\"")

    print()
    print("Building answer key and study data ...")
    answer_key_rows, study_questions = build_study(
        test_records, long_df, sampled_idx, letter_assignments
    )

    print()
    write_answer_key(answer_key_rows, OUT_DIR / "answer_key.csv")

    print()
    generate_workbook(study_questions, OUT_DIR / "questionnaire.xlsx")

    print()
    print("-" * 60)
    print("Summary")
    print(f"  Questions sampled      : {len(sampled_idx)}")
    print(f"  Answer-key rows        : {len(answer_key_rows)}  (50 answers)")
    print(f"  Score cells in workbook: {len(sampled_idx) * 5 * 4}"
          f"  (10 questions x 5 answers x 4 dimensions)")
    print()
    print("REMINDER: answer_key.csv is PRIVATE - do not share with raters.")
    print("          Distribute questionnaire.xlsx; collect filled copies by email.")


if __name__ == "__main__":
    main()
